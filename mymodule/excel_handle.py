import logging
from pathlib import Path
# import json
from typing import Dict, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelHandler:
    def __init__(self, excel_path: Path):
        """初始化Excel处理器"""
        self.current_time = "2025-05-13 05:46:27"

        self.excel_path = Path(excel_path)
        self.title_checked = False

        # 定义列映射
        self.COLUMN_MAPPING = {
            'student_id': 'I',  # 学号列
            'thesis_title': 'U',  # 论文题目列
            'thesis_file': 'Y',  # 论文文件名列
            'support_file': 'Z',  # 支撑材料文件名列
            'report_file': 'AA'  # 查重报告文件名列
        }

        try:
            # 加载Excel工作簿
            self.workbook = load_workbook(filename=str(self.excel_path))
            self.sheet = self.workbook.active

            # 验证所需列是否存在
            max_col = self.sheet.max_column
            if column_index_from_string('AA') > max_col:
                raise ValueError(f"Excel文件列数不足，需要到AA列，当前只有{get_column_letter(max_col)}列")

            logging.info(f"成功加载Excel文件")

        except Exception as e:
            logging.error(f"加载Excel文件失败: {str(e)}")
            raise

    def _get_cell_value(self, row: int, col: str) -> str:
        """获取单元格的值"""
        value = self.sheet[f"{col}{row}"].value
        return str(value).strip() if value is not None else ""

    def _set_cell_value(self, row: int, col: str, value: str) -> None:
        """设置单元格的值"""
        if isinstance(value, Path):
            value = value.name  # 只使用文件名部分
        self.sheet[f"{col}{row}"] = str(value)  # 确保值是字符串

    def _find_student_row(self, student_id: str) -> Optional[int]:
        """在Excel中查找学生所在行"""
        student_id = str(student_id).strip()
        col = self.COLUMN_MAPPING['student_id']

        # 从第2行开始搜索（第1行是标题）
        for row in range(2, self.sheet.max_row + 1):
            current_id = self._get_cell_value(row, col)
            if current_id == student_id:
                logging.info(f"找到学生 {student_id} 在第 {row} 行")
                return row

        logging.warning(f"未找到学号为 {student_id} 的学生")
        return None

    def _compare_titles(self, row: int, json_title: str) -> bool:
        """比对论文题目是否一致"""
        excel_title = self._get_cell_value(row, self.COLUMN_MAPPING['thesis_title'])
        json_title = str(json_title).strip()

        # 清理标题
        excel_title = ''.join(excel_title.split())
        json_title = ''.join(json_title.split())

        if excel_title.lower() != json_title.lower():
            logging.warning(
                f"题目不一致:\nExcel: {excel_title}\nJSON: {json_title}"
            )
            return False

        return True

    def strict_title_check(self, student_id: str, json_path: Path) -> bool:

        global json
        if self.title_checked:
            # 如果不是第一次比对，只返回比对结果
            row = self._find_student_row(student_id)
            if row is None:
                return False

            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    import json
                    json_data = json.load(f)
                    json_title = json_data.get('title', '')
                return self._compare_titles(row, json_title)
            except Exception as e:
                logging.error(f"读取JSON文件失败: {str(e)}")
                return False
        else:
            # 第一次比对，失败就抛出异常
            row = self._find_student_row(student_id)
            if row is None:
                raise ValueError(f"未找到学号为 {student_id} 的学生")

            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    import json
                    json_data = json.load(f)
                    json_title = json_data.get('title', '')

                if not self._compare_titles(row, json_title):
                    raise ValueError(
                        f"学号 {student_id} 的论文题目不匹配\n"
                        f"Excel中的题目: {self._get_cell_value(row, self.COLUMN_MAPPING['thesis_title'])}\n"
                        f"JSON中的题目: {json_title}"
                    )

                # 标记已经进行过题目比对
                self.title_checked = True
                return True

            except json.JSONDecodeError as e:
                raise ValueError(f"JSON文件格式错误: {str(e)}")
            except Exception as e:
                raise ValueError(f"读取识别结果失败: {str(e)}")

    def process_student(self, student_id: str, file_names: Dict[str, str], json_path: Path = None) -> Tuple[bool, str]:

        # 处理单个学生的所有文件
        try:
            messages = []

            # 查找学生
            row = self._find_student_row(student_id)
            if row is None:
                return False, f"未找到学号为 {student_id} 的学生"

            # 如果是论文或支撑材料，需要进行题目比对
            if ('thesis' in file_names or 'ktbg' in file_names) and json_path is not None:
                try:
                    if not self.strict_title_check(student_id, json_path):
                        return False, f"学号 {student_id} 的论文题目不匹配"
                except ValueError as e:
                    # 第一次比对失败，直接抛出异常终止程序
                    raise

            # 更新文件名
            type_to_column = {
                'thesis': 'thesis_file',
                'report': 'report_file',
                'ktbg': 'support_file'
            }

            for file_type, file_path in file_names.items():
                if file_type in type_to_column:
                    # 如果是Path对象，获取文件名；如果是字符串，直接使用
                    file_name = Path(file_path).name if isinstance(file_path, (Path, str)) else str(file_path)

                    col = self.COLUMN_MAPPING[type_to_column[file_type]]
                    self._set_cell_value(row, col, file_name)
                    messages.append(f"{file_type}文件已更新")

            # 保存Excel
            self.workbook.save(str(self.excel_path))
            messages.append("Excel文件已保存")

            return True, "; ".join(messages)

        except Exception as e:
            error_msg = f"处理学生 {student_id} 的文件时出错: {str(e)}"
            logging.error(error_msg)
            return False, error_msg

    def __del__(self):
        """析构函数，确保工作簿被关闭"""
        try:
            self.workbook.close()
        except:
            pass